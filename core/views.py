from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth import get_user_model
from django.db.models import Count, Q, Avg
from django.utils import timezone
from django.contrib import messages
from datetime import datetime, timedelta
from happycall.models import HappyCall
from services.models import ServiceRequest
from customers.models import Customer
from employees.models import Employee
from scheduling.models import Schedule
from .forms import DataUploadForm
from .upload_handlers import DataUploadHandler

User = get_user_model()


@login_required
def dashboard(request):
    """
    메인 대시보드 페이지
    관리자는 관리 대시보드, 일반 직원은 기본 대시보드
    """
    user = request.user

    # 관리자인 경우 관리 대시보드로 리다이렉트
    if user.is_superuser:
        return admin_dashboard(request)

    context = {
        "user": user,
        "is_superuser": user.is_superuser,
        "user_groups": user.groups.all(),
    }

    return render(request, "core/dashboard.html", context)


@login_required
def admin_dashboard(request):
    """관리자 대시보드"""
    if not request.user.is_superuser:
        return dashboard(request)

    # 기본 통계
    today = timezone.now().date()
    week_ago = today - timedelta(days=7)
    month_ago = today - timedelta(days=30)

    # 전체 현황
    total_customers = Customer.objects.count()
    total_services = ServiceRequest.objects.count()
    total_happycalls = HappyCall.objects.count()
    total_employees = Employee.objects.filter(status="active").count()

    # 이번 주/월 신규
    week_customers = Customer.objects.filter(created_at__gte=week_ago).count()
    month_services = ServiceRequest.objects.filter(created_at__gte=month_ago).count()
    week_happycalls = HappyCall.objects.filter(created_at__gte=week_ago).count()

    # 직원별 해피콜 배정 현황
    employee_happycall_stats = []
    employees = Employee.objects.filter(status="active").select_related("user")

    for employee in employees:
        # 진행 중인 해피콜 (미완료)
        pending_calls = (
            HappyCall.objects.filter(
                Q(
                    first_call_caller=employee.user,
                    call_stage__in=["1st_pending", "1st_in_progress"],
                )
                | Q(
                    second_call_caller=employee.user,
                    call_stage__in=["2nd_pending", "2nd_in_progress"],
                )
                | Q(
                    third_call_caller=employee.user,
                    call_stage__in=["3rd_pending", "3rd_in_progress"],
                )
                | Q(
                    fourth_call_caller=employee.user,
                    call_stage__in=["4th_pending", "4th_in_progress"],
                )
            )
            .distinct()
            .count()
        )

        # 완료된 해피콜 (이번 주)
        completed_calls = (
            HappyCall.objects.filter(
                Q(
                    first_call_caller=employee.user,
                    first_call_success=True,
                    first_call_date__gte=week_ago,
                )
                | Q(
                    second_call_caller=employee.user,
                    second_call_success=True,
                    second_call_date__gte=week_ago,
                )
                | Q(
                    third_call_caller=employee.user,
                    third_call_success=True,
                    third_call_date__gte=week_ago,
                )
                | Q(
                    fourth_call_caller=employee.user,
                    fourth_call_success=True,
                    fourth_call_date__gte=week_ago,
                )
            )
            .distinct()
            .count()
        )

        # 전체 담당 해피콜
        total_assigned = (
            HappyCall.objects.filter(
                Q(first_call_caller=employee.user)
                | Q(second_call_caller=employee.user)
                | Q(third_call_caller=employee.user)
                | Q(fourth_call_caller=employee.user)
            )
            .distinct()
            .count()
        )

        employee_happycall_stats.append(
            {
                "employee": employee,
                "pending_calls": pending_calls,
                "completed_calls_week": completed_calls,
                "total_assigned": total_assigned,
            }
        )

    # 해피콜 단계별 현황
    happycall_stage_stats = {
        "1st_pending": HappyCall.objects.filter(call_stage="1st_pending").count(),
        "1st_completed": HappyCall.objects.filter(call_stage="1st_completed").count(),
        "2nd_pending": HappyCall.objects.filter(call_stage="2nd_pending").count(),
        "2nd_completed": HappyCall.objects.filter(call_stage="2nd_completed").count(),
        "3rd_pending": HappyCall.objects.filter(call_stage="3rd_pending").count(),
        "3rd_completed": HappyCall.objects.filter(call_stage="3rd_completed").count(),
        "4th_pending": HappyCall.objects.filter(call_stage="4th_pending").count(),
        "completed": HappyCall.objects.filter(call_stage="completed").count(),
    }

    # 서비스 타입별 통계 (이번 달)
    service_type_stats = (
        ServiceRequest.objects.filter(created_at__gte=month_ago)
        .values("service_type__name")
        .annotate(count=Count("id"))
        .order_by("-count")
    )

    # 최근 일정 (오늘부터 일주일)
    upcoming_schedules = (
        Schedule.objects.filter(
            start_datetime__gte=timezone.now(),
            start_datetime__lte=timezone.now() + timedelta(days=7),
        )
        .select_related("assignee", "department")
        .order_by("start_datetime")[:10]
    )

    context = {
        "is_admin_dashboard": True,
        "total_customers": total_customers,
        "total_services": total_services,
        "total_happycalls": total_happycalls,
        "total_employees": total_employees,
        "week_customers": week_customers,
        "month_services": month_services,
        "week_happycalls": week_happycalls,
        "employee_happycall_stats": employee_happycall_stats,
        "happycall_stage_stats": happycall_stage_stats,
        "service_type_stats": service_type_stats,
        "upcoming_schedules": upcoming_schedules,
        "today": today,
    }

    return render(request, "core/admin_dashboard.html", context)


@login_required
def data_upload(request):
    """데이터 업로드 페이지"""
    if not request.user.is_superuser:
        messages.error(request, "관리자만 접근할 수 있습니다.")
        return redirect("core:dashboard")

    if request.method == "POST":
        form = DataUploadForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                # 파일 처리
                df = form.process_file()

                # progress_key 받아오기 (POST 데이터 또는 새로 생성)
                progress_key = request.POST.get("progress_key")
                if not progress_key:
                    import uuid

                    progress_key = str(uuid.uuid4())

                handler = DataUploadHandler(
                    upload_type=form.cleaned_data["upload_type"],
                    duplicate_handling=form.cleaned_data["duplicate_handling"],
                    user=request.user,
                    progress_key=progress_key,
                )

                results = handler.process_data(df)

                # 최종 진행률 완료 표시
                handler.update_progress(100, "업로드 완료!")

                # AJAX 요청인 경우 JSON 응답
                if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                    from django.http import JsonResponse

                    return JsonResponse(
                        {
                            "success": True,
                            "progress_key": progress_key,
                            "results": results,
                        }
                    )

                # 결과 메시지
                success_msg = f"성공: {results['success']}건"
                if results["updated"] > 0:
                    success_msg += f", 업데이트: {results['updated']}건"
                if results["skipped"] > 0:
                    success_msg += f", 건너뜀: {results['skipped']}건"

                messages.success(request, f"업로드 완료! {success_msg}")

                if results["errors"] > 0:
                    error_msg = f"오류: {results['errors']}건"
                    for error in results["error_details"][:5]:  # 최대 5개만 표시
                        error_msg += f"\\n- {error}"
                    if len(results["error_details"]) > 5:
                        error_msg += f"\\n... 외 {len(results['error_details']) - 5}건"
                    messages.warning(request, error_msg)

                return redirect("core:data_upload")

            except Exception as e:
                # AJAX 요청인 경우 JSON 오류 응답
                if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                    from django.http import JsonResponse

                    return JsonResponse({"success": False, "error": str(e)})
                messages.error(request, f"파일 처리 중 오류가 발생했습니다: {str(e)}")
    else:
        form = DataUploadForm()

    # Vehicle 모델 import
    from customers.models import Vehicle

    # 최근 업로드 통계
    upload_stats = {
        "total_customers": Customer.objects.count(),
        "total_vehicles": Vehicle.objects.count(),
        "total_services": ServiceRequest.objects.count(),
    }

    context = {
        "form": form,
        "upload_stats": upload_stats,
    }

    return render(request, "core/data_upload.html", context)


@login_required
def download_template(request, template_type):
    """데이터 업로드 템플릿 다운로드"""
    if not request.user.is_superuser:
        messages.error(request, "관리자만 접근할 수 있습니다.")
        return redirect("core:dashboard")

    from django.http import HttpResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    # 새 워크북 생성
    wb = openpyxl.Workbook()
    ws = wb.active

    # 템플릿별 헤더 정의
    templates = {
        "customers": {
            "filename": "customer_data_upload_template.xlsx",
            "sheet_name": "고객 데이터",
            "headers": [
                "name",
                "phone",
                "email",
                "address_main",
                "address_detail",
                "customer_type",
                "membership_status",
                "customer_grade",
                "business_number",
                "company_name",
                "privacy_consent",
                "marketing_consent",
                "do_not_contact",
                "is_banned",
                "notes",
            ],
            "sample_data": [
                [
                    "홍길동",
                    "010-1234-5678",
                    "hong@example.com",
                    "서울시 강남구",
                    "테헤란로 123",
                    "individual",
                    "basic",
                    "A",
                    "",
                    "",
                    "TRUE",
                    "FALSE",
                    "FALSE",
                    "FALSE",
                    "샘플 고객 데이터",
                ]
            ],
        },
        "vehicles": {
            "filename": "vehicle_data_upload_template.xlsx",
            "sheet_name": "차량 데이터",
            "headers": [
                "vehicle_number",
                "model",
                "year",
                "model_detail",
                "mileage",
                "customer_phone",
                "vehicle_type",
                "fuel_type",
                "notes",
            ],
            "sample_data": [
                [
                    "123가4567",
                    "소나타",
                    "2020",
                    "하이브리드",
                    "50000",
                    "010-1234-5678",
                    "sedan",
                    "gasoline",
                    "샘플 차량 데이터",
                ]
            ],
        },
        "services": {
            "filename": "service_data_upload_template.xlsx",
            "sheet_name": "서비스 데이터",
            "headers": [
                "customer_phone",
                "vehicle_number",
                "service_type",
                "service_date",
                "status",
                "priority",
                "description",
                "estimated_price",
                "notes",
            ],
            "sample_data": [
                [
                    "010-1234-5678",
                    "123가4567",
                    "engine_oil",
                    "2024-01-15",
                    "completed",
                    "normal",
                    "엔진오일 교환",
                    "50000",
                    "샘플 서비스 데이터",
                ]
            ],
        },
    }

    if template_type not in templates:
        messages.error(request, "잘못된 템플릿 타입입니다.")
        return redirect("core:data_upload")

    template = templates[template_type]
    ws.title = template["sheet_name"]

    # 헤더 스타일링
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )

    # 헤더 작성
    for col, header in enumerate(template["headers"], 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # 샘플 데이터 작성
    for row_idx, sample_row in enumerate(template["sample_data"], 2):
        for col_idx, value in enumerate(sample_row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # 컬럼 너비 자동 조정
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # HTTP 응답 생성
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    from urllib.parse import quote

    filename = template["filename"]
    response["Content-Disposition"] = (
        f"attachment; filename=\"{filename}\"; filename*=UTF-8''{quote(filename)}"
    )

    # 워크북을 응답에 저장
    wb.save(response)

    return response


@login_required
def upload_progress(request, progress_key):
    """업로드 진행 상황 조회 API"""
    if not request.user.is_superuser:
        return JsonResponse({"error": "권한이 없습니다."}, status=403)

    from django.http import JsonResponse
    from django.core.cache import cache

    progress_data = cache.get(f"upload_progress_{progress_key}")

    if progress_data:
        return JsonResponse(progress_data)
    else:
        return JsonResponse(
            {
                "percent": 0,
                "message": "진행 상황을 찾을 수 없습니다.",
                "results": {
                    "success": 0,
                    "updated": 0,
                    "skipped": 0,
                    "errors": 0,
                    "error_details": [],
                },
            }
        )


@login_required
def start_upload(request):
    """업로드 시작 - progress_key 생성"""
    if not request.user.is_superuser:
        return JsonResponse({"error": "권한이 없습니다."}, status=403)

    from django.http import JsonResponse
    import uuid

    # progress_key 생성
    progress_key = str(uuid.uuid4())

    # 초기 진행 상태 설정
    from django.core.cache import cache

    progress_data = {
        "percent": 0,
        "message": "업로드 준비 중...",
        "results": {
            "success": 0,
            "updated": 0,
            "skipped": 0,
            "errors": 0,
            "error_details": [],
        },
    }
    cache.set(f"upload_progress_{progress_key}", progress_data, timeout=300)

    return JsonResponse({"success": True, "progress_key": progress_key})
