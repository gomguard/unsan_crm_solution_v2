from django.shortcuts import render, get_object_or_404, redirect
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.urls import reverse_lazy
from django.http import JsonResponse, HttpResponseRedirect, HttpResponse
from django.db.models import Q
from django.core.paginator import Paginator
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

from .models import Customer


class CustomerListView(LoginRequiredMixin, ListView):
    model = Customer
    template_name = 'customers/customer_list.html'
    context_object_name = 'customers'
    paginate_by = 20
    
    def get(self, request, *args, **kwargs):
        # 엑셀 내보내기 요청 처리
        if request.GET.get('export') == 'excel':
            return self.export_to_excel()
        return super().get(request, *args, **kwargs)
    
    def post(self, request, *args, **kwargs):
        # 엑셀 내보내기 요청 처리 (POST)
        if request.POST.get('export') == 'excel':
            return self.export_to_excel()
        return super().get(request, *args, **kwargs)
    
    def get_queryset(self):
        queryset = Customer.objects.filter(is_active=True).prefetch_related('vehicle_ownerships')
        
        # 검색 기능
        search_query = self.request.GET.get('search')
        if search_query:
            queryset = queryset.filter(
                Q(name__icontains=search_query) |
                Q(phone__icontains=search_query) |
                Q(email__icontains=search_query) |
                Q(company_name__icontains=search_query)
            )
        
        # 고객 구분 필터
        customer_type = self.request.GET.get('customer_type')
        if customer_type:
            queryset = queryset.filter(customer_type=customer_type)
        
        # 멤버십 상태 필터
        membership_status = self.request.GET.get('membership_status')
        if membership_status:
            queryset = queryset.filter(membership_status=membership_status)
        
        # 고객 등급 필터
        customer_grade = self.request.GET.get('customer_grade')
        if customer_grade:
            queryset = queryset.filter(customer_grade=customer_grade)
        
        # 서비스 횟수 필터
        service_count_min = self.request.GET.get('service_count_min')
        service_count_max = self.request.GET.get('service_count_max')
        if service_count_min:
            try:
                queryset = queryset.filter(total_service_count__gte=int(service_count_min))
            except ValueError:
                pass
        if service_count_max:
            try:
                queryset = queryset.filter(total_service_count__lte=int(service_count_max))
            except ValueError:
                pass
        
        # 사용 금액 필터 (만원 단위로 입력받음)
        service_amount_min = self.request.GET.get('service_amount_min')
        service_amount_max = self.request.GET.get('service_amount_max')
        if service_amount_min:
            try:
                # 만원 단위를 원 단위로 변환
                min_amount = int(service_amount_min) * 10000
                queryset = queryset.filter(total_service_amount__gte=min_amount)
            except ValueError:
                pass
        if service_amount_max:
            try:
                # 만원 단위를 원 단위로 변환
                max_amount = int(service_amount_max) * 10000
                queryset = queryset.filter(total_service_amount__lte=max_amount)
            except ValueError:
                pass
        
        # 정렬
        ordering = self.request.GET.get('ordering', '-created_at')
        if ordering:
            queryset = queryset.order_by(ordering)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # 필터링된 고객 수 계산
        filtered_customers = self.get_queryset()
        filtered_count = filtered_customers.count()
        
        # 고객별 전화번호 처리
        customers_with_phone = []
        # 페이지네이션된 객체 사용
        page_customers = context.get('page_obj', context.get('customers', []))
        
        # 관리자의 전화번호 전체보기 옵션 확인
        show_full_phone = self.request.GET.get('show_full_phone') == '1'
        
        for customer in page_customers:
            # 모든 전화번호 처리를 Customer 모델의 메서드로 위임
            display_phone = customer.get_phone_for_user(self.request.user, show_full_phone)
            
            customers_with_phone.append({
                'customer': customer,
                'display_phone': display_phone,
                'can_view_phone': customer.can_view_phone(self.request.user, show_full_phone)
            })
        
        context.update({
            'customers_with_phone': customers_with_phone,
            'show_full_phone': show_full_phone,
            'is_admin': self.request.user.is_superuser,
            'search_query': self.request.GET.get('search', ''),
            'customer_type_filter': self.request.GET.get('customer_type', ''),
            'membership_status_filter': self.request.GET.get('membership_status', ''),
            'customer_grade_filter': self.request.GET.get('customer_grade', ''),
            'service_count_min': self.request.GET.get('service_count_min', ''),
            'service_count_max': self.request.GET.get('service_count_max', ''),
            'service_amount_min': self.request.GET.get('service_amount_min', ''),
            'service_amount_max': self.request.GET.get('service_amount_max', ''),
            'ordering': self.request.GET.get('ordering', '-created_at'),
            'total_customers': Customer.objects.filter(is_active=True).count(),
            'filtered_customers': filtered_count,
            'vip_customers': Customer.objects.filter(is_active=True, membership_status='vip').count(),
            'corporate_customers': Customer.objects.filter(is_active=True, customer_type='corporate').count(),
        })
        return context
    
    def export_to_excel(self):
        """고객 목록을 엑셀로 내보내기"""
        # 동일한 필터링 로직 적용 (페이지네이션 제외)
        queryset = self.get_queryset()
        
        # 엑셀 워크북 생성
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "고객목록"
        
        # 헤더 스타일 설정
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 헤더 추가
        headers = [
            "이름", "전화번호", "이메일", "주소", "멤버십상태", "고객등급",
            "차량대수", "서비스횟수", "사용금액", "최근방문일", "등록일"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 데이터 추가
        show_full_phone = self.request.GET.get('show_full_phone') == '1'
        for row_num, customer in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=customer.name or "미등록")
            # 엑셀 내보내기에서도 전화번호 보안 적용
            ws.cell(row=row_num, column=2, value=customer.get_phone_for_user(self.request.user, show_full_phone))
            ws.cell(row=row_num, column=3, value=customer.email or "")
            ws.cell(row=row_num, column=4, value=customer.get_full_address())
            ws.cell(row=row_num, column=5, value=customer.get_membership_status_display())
            ws.cell(row=row_num, column=6, value=f"{customer.customer_grade}등급" if customer.customer_grade else "")
            ws.cell(row=row_num, column=7, value=f"{customer.vehicle_ownerships.count()}대")
            ws.cell(row=row_num, column=8, value=f"{customer.total_service_count}회")
            ws.cell(row=row_num, column=9, value=f"{customer.total_service_amount:,.0f}원")
            ws.cell(row=row_num, column=10, value=customer.last_service_date.strftime("%Y.%m.%d") if customer.last_service_date else "-")
            ws.cell(row=row_num, column=11, value=customer.created_at.strftime("%Y.%m.%d"))
        
        # 열 너비 자동 조정
        column_widths = [12, 15, 25, 30, 12, 10, 10, 12, 15, 12, 12]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # 파일명 생성 - 필터 조건 반영
        filename_parts = ["고객목록"]
        
        # 검색 조건 추가
        search_query = self.request.GET.get('search')
        if search_query:
            filename_parts.append(f"검색_{search_query[:10]}")
        
        # 고객 구분 추가
        customer_type = self.request.GET.get('customer_type')
        if customer_type == 'individual':
            filename_parts.append("개인")
        elif customer_type == 'corporate':
            filename_parts.append("법인")
        
        # 멤버십 상태 추가
        membership_status = self.request.GET.get('membership_status')
        if membership_status == 'vip':
            filename_parts.append("VIP")
        elif membership_status == 'premium':
            filename_parts.append("우수회원")
        elif membership_status == 'basic':
            filename_parts.append("일반회원")
        elif membership_status == 'none':
            filename_parts.append("비회원")
        
        # 등급 추가
        customer_grade = self.request.GET.get('customer_grade')
        if customer_grade:
            filename_parts.append(f"{customer_grade}등급")
        
        # 서비스 횟수 범위 추가
        service_count_min = self.request.GET.get('service_count_min')
        service_count_max = self.request.GET.get('service_count_max')
        if service_count_min or service_count_max:
            count_range = f"서비스{service_count_min or '0'}회~{service_count_max or '무제한'}회"
            filename_parts.append(count_range)
        
        # 사용금액 범위 추가
        service_amount_min = self.request.GET.get('service_amount_min')
        service_amount_max = self.request.GET.get('service_amount_max')
        if service_amount_min or service_amount_max:
            amount_range = f"금액{service_amount_min or '0'}만~{service_amount_max or '무제한'}만"
            filename_parts.append(amount_range)
        
        # 데이터 건수 추가
        filename_parts.append(f"{queryset.count()}건")
        
        # 최종 파일명 생성
        filename_base = "_".join(filename_parts)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{filename_base}_{timestamp}.xlsx"
        
        # HTTP 응답 생성
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # 엑셀 파일 저장
        wb.save(response)
        return response


class CustomerDetailView(LoginRequiredMixin, DetailView):
    model = Customer
    template_name = 'customers/customer_detail.html'
    context_object_name = 'customer'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        customer = self.get_object()
        
        # 서비스 이력 가져오기 (최근 10개) - 차량 정보 포함
        from services.models import ServiceRequest
        from django.db.models import Q
        
        services = ServiceRequest.objects.filter(
            Q(customer=customer) | 
            Q(temp_customer_name=customer.name, temp_customer_phone=customer.phone)
        ).select_related('assigned_employee', 'vehicle').order_by('-created_at')[:10]
        
        # 통계 계산
        all_services = ServiceRequest.objects.filter(
            Q(customer=customer) | 
            Q(temp_customer_name=customer.name, temp_customer_phone=customer.phone)
        )
        
        service_count = all_services.count()
        total_payment = sum(service.estimated_price or 0 for service in all_services if service.estimated_price)
        last_service = all_services.first()
        last_service_date = last_service.created_at if last_service else None
        
        # 고객 차량 정보 가져오기
        vehicles = customer.vehicle_ownerships.filter(end_date__isnull=True).select_related('vehicle')
        
        context.update({
            'services': services,
            'service_count': service_count,
            'total_payment': total_payment,
            'last_service_date': last_service_date,
            'vehicles': vehicles,
        })
        
        return context


class CustomerCreateView(LoginRequiredMixin, CreateView):
    model = Customer
    template_name = 'customers/customer_form.html'
    fields = [
        'customer_type', 'name', 'phone', 'email',
        'address_main', 'address_detail',
        'business_number', 'company_name',
        'membership_status', 'customer_grade', 'notes'
    ]
    
    def form_valid(self, form):
        messages.success(self.request, f'{form.instance.name} 고객이 성공적으로 등록되었습니다.')
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form_title'] = '고객 등록'
        context['form_action'] = 'create'
        return context


class CustomerUpdateView(LoginRequiredMixin, UpdateView):
    model = Customer
    template_name = 'customers/customer_form.html'
    fields = [
        'customer_type', 'name', 'phone', 'email',
        'address_main', 'address_detail',
        'business_number', 'company_name',
        'membership_status', 'customer_grade', 'notes'
    ]
    
    def form_valid(self, form):
        messages.success(self.request, f'{form.instance.name} 고객 정보가 성공적으로 수정되었습니다.')
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form_title'] = '고객 정보 수정'
        context['form_action'] = 'edit'
        return context


class CustomerDeleteView(LoginRequiredMixin, DeleteView):
    model = Customer
    template_name = 'customers/customer_confirm_delete.html'
    success_url = reverse_lazy('customers:customer_list')
    
    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        success_url = self.get_success_url()
        
        # 실제로는 삭제하지 않고 비활성화
        self.object.is_active = False
        self.object.save()
        
        messages.success(request, f'{self.object.name} 고객이 비활성화되었습니다.')
        return HttpResponseRedirect(success_url)


@login_required
def customer_search(request):
    """AJAX 고객 검색"""
    query = request.GET.get('q', '')
    
    if len(query) < 2:
        return JsonResponse({'customers': []})
    
    customers = Customer.objects.filter(
        Q(name__icontains=query) |
        Q(phone__icontains=query),
        is_active=True
    )[:10]
    
    customer_data = []
    for customer in customers:
        customer_data.append({
            'id': customer.id,
            'name': customer.name,
            'phone': customer.phone,
            'customer_type': customer.get_customer_type_display(),
            'membership_status': customer.get_membership_status_display(),
            'url': customer.get_absolute_url(),
        })
    
    return JsonResponse({'customers': customer_data})


@login_required
def toggle_customer_active(request, pk):
    """고객 활성화/비활성화 토글"""
    customer = get_object_or_404(Customer, pk=pk)
    
    if request.method == 'POST':
        customer.is_active = not customer.is_active
        customer.save()
        
        status = '활성화' if customer.is_active else '비활성화'
        messages.success(request, f'{customer.name} 고객이 {status}되었습니다.')
    
    return redirect('customers:customer_detail', pk=pk)
